from openpyxl import Workbook
import ConfigParser
import Config


class ExcelDoc:

    def __init__(self):
        pass

    def create_excel_doc(self, info, monday, tuesday, wednesday,
                         thursday, friday, saturday, sunday):

        workbook = Workbook()
        sheet1 = workbook.active
        sheet1.title = info
        sheet2 = workbook.create_sheet(monday)
        sheet3 = workbook.create_sheet(tuesday)
        sheet4 = workbook.create_sheet(wednesday)
        sheet5 = workbook.create_sheet(thursday)
        sheet6 = workbook.create_sheet(friday)
        sheet7 = workbook.create_sheet(saturday)
        sheet8 = workbook.create_sheet(sunday)

        return workbook

    def fill_excel_doc(self, config_path):

        config = ConfigParser.ConfigParser()
        config.read(config_path)

        config_update = Config.Config()
        config_update.update_week(config_path)

        workbook = self.create_excel_doc('Info', 'Monday', 'Tuesday', 'Wednesday',
                                         'Thursday', 'Friday', 'Saturday', 'Sunday')
        info = workbook.active

        for sheet in workbook:
            sheet['A1'] = 'Entrepeneur'
            sheet['B1'] = 'Description of work done'
            sheet['C1'] = 'Amount of hours'

        info['A1'] = 'This week is: ' + config.get('General', 'weeknumber')
        info['A2'] = 'Employee: ' + config.get('UserInfo', 'username')
        info['B1'] = None
        info['C1'] = None

        workbook.save('weeknumber.xlsx')

        return
